"""Écriture des demandes dans le fichier Excel de suivi."""

from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import Workbook

HEADERS = ["Date", "Nom", "Prénom", "Email", "Téléphone", "Scolarité", "Brochure(s)", "Statut"]


def _ensure_workbook(path: Path, sheet_name: str) -> openpyxl.Workbook:
    """Charge ou crée le classeur et s'assure que l'onglet existe avec les en-têtes."""
    if path.exists():
        wb = openpyxl.load_workbook(path)
    else:
        wb = Workbook()
        # Supprimer la feuille par défaut si elle existe
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append(HEADERS)
    return wb


def log_demande(
    excel_path: str,
    sheet_name: str,
    nom: str,
    prenom: str,
    email: str,
    telephone: str,
    scolarite: str,
    brochures: list[str],
    statut: str,
) -> None:
    """
    Ajoute une ligne dans l'Excel.
    Lève une exception si le fichier est verrouillé ou inaccessible.
    """
    path = Path(excel_path)

    # Vérifier si le fichier est verrouillé (fichier temporaire Office)
    lock_file = path.parent / f"~${path.name}"
    if lock_file.exists():
        raise PermissionError(
            f"Le fichier Excel est ouvert dans une autre application. "
            f"Ferme '{path.name}' puis relance le traitement."
        )

    wb = _ensure_workbook(path, sheet_name)
    ws = wb[sheet_name]

    row = [
        datetime.now().strftime("%d/%m/%Y %H:%M"),
        nom,
        prenom,
        email,
        telephone,
        scolarite,
        ", ".join(brochures),
        statut,
    ]
    ws.append(row)

    # Ajuster la largeur des colonnes (best-effort)
    try:
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
    except Exception:
        pass

    wb.save(path)
