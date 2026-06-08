"""
Application principale MMPP Brochures.
Interface graphique Tkinter — un bouton + zone de log.
"""

import json
import threading
import tkinter as tk
from pathlib import Path
from tkinter import scrolledtext, ttk

from brevo_client import send_brochures
from excel_logger import log_demande
from gmail_client import (
    fetch_unprocessed_messages,
    get_message_body,
    get_or_create_label,
    get_service,
    mark_as_processed,
)
from parser import parse_body

CONFIG_PATH = Path("config.json")


def load_config() -> dict:
    if not CONFIG_PATH.exists():
        raise FileNotFoundError(
            "config.json introuvable. Copie config.example.json → config.json et remplis les valeurs."
        )
    with open(CONFIG_PATH) as f:
        return json.load(f)


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("MMPP Brochures — Traitement automatique")
        self.geometry("800x540")
        self.resizable(True, True)
        self.configure(bg="#f0f0f0")

        self._build_ui()
        self._load_config_safe()

    def _build_ui(self):
        top = tk.Frame(self, bg="#f0f0f0", pady=10)
        top.pack(fill="x", padx=16)

        self.btn_process = tk.Button(
            top,
            text="▶  Traiter les nouvelles demandes",
            font=("Helvetica", 14, "bold"),
            bg="#1a73e8",
            fg="white",
            activebackground="#1558b0",
            activeforeground="white",
            padx=20,
            pady=8,
            relief="flat",
            cursor="hand2",
            command=self._on_process_click,
        )
        self.btn_process.pack(side="left")

        self.dry_var = tk.BooleanVar()
        self.chk_dry = tk.Checkbutton(
            top,
            text="Mode dry-run (pas d'envoi réel)",
            variable=self.dry_var,
            bg="#f0f0f0",
            font=("Helvetica", 12),
        )
        self.chk_dry.pack(side="left", padx=20)

        self.status_lbl = tk.Label(top, text="", bg="#f0f0f0", font=("Helvetica", 11, "italic"), fg="#555")
        self.status_lbl.pack(side="left")

        # Zone de log
        log_frame = tk.Frame(self, bg="#f0f0f0")
        log_frame.pack(fill="both", expand=True, padx=16, pady=(0, 12))

        tk.Label(log_frame, text="Journal", bg="#f0f0f0", font=("Helvetica", 11, "bold")).pack(anchor="w")
        self.log_area = scrolledtext.ScrolledText(
            log_frame,
            font=("Menlo", 11),
            bg="#1e1e1e",
            fg="#d4d4d4",
            insertbackground="white",
            state="disabled",
            wrap="word",
        )
        self.log_area.pack(fill="both", expand=True)

        # Tags de couleur pour le log
        self.log_area.tag_config("ok", foreground="#6fcf97")
        self.log_area.tag_config("warn", foreground="#f2c94c")
        self.log_area.tag_config("err", foreground="#eb5757")
        self.log_area.tag_config("info", foreground="#90caf9")

        # Bouton effacer
        tk.Button(
            self,
            text="Effacer le journal",
            bg="#e0e0e0",
            relief="flat",
            command=self._clear_log,
        ).pack(anchor="e", padx=16, pady=(0, 8))

    def _load_config_safe(self):
        try:
            cfg = load_config()
            if cfg.get("dry_run"):
                self.dry_var.set(True)
            self._log("Config chargée. Prêt.", "ok")
        except FileNotFoundError as e:
            self._log(str(e), "err")

    def _log(self, msg: str, tag: str = "info"):
        self.log_area.config(state="normal")
        self.log_area.insert("end", msg + "\n", tag)
        self.log_area.see("end")
        self.log_area.config(state="disabled")

    def _clear_log(self):
        self.log_area.config(state="normal")
        self.log_area.delete("1.0", "end")
        self.log_area.config(state="disabled")

    def _on_process_click(self):
        self.btn_process.config(state="disabled", text="⏳  Traitement en cours…")
        self.status_lbl.config(text="")
        t = threading.Thread(target=self._run_processing, daemon=True)
        t.start()

    def _run_processing(self):
        try:
            cfg = load_config()
        except FileNotFoundError as e:
            self._log(str(e), "err")
            self._finish()
            return

        dry_run = self.dry_var.get() or cfg.get("dry_run", False)
        if dry_run:
            self._log("⚠️  MODE DRY-RUN — aucun mail ne sera envoyé.", "warn")

        try:
            self._log("Connexion à Gmail…", "info")
            service = get_service()
            label_id = get_or_create_label(service, cfg["processed_label"])
            self._log(f"Label « {cfg['processed_label']} » : {label_id}", "info")
        except Exception as e:
            self._log(f"Erreur Gmail : {e}", "err")
            self._finish()
            return

        try:
            messages = fetch_unprocessed_messages(service, cfg["gmail_query"], cfg["processed_label"])
        except Exception as e:
            self._log(f"Erreur lecture Gmail : {e}", "err")
            self._finish()
            return

        if not messages:
            self._log("Aucun nouveau mail à traiter.", "ok")
            self._finish()
            return

        self._log(f"{len(messages)} mail(s) trouvé(s).", "info")

        ok_count = 0
        manual_count = 0
        error_count = 0

        for msg in messages:
            msg_id = msg["id"]
            self._log(f"\n── Mail {msg_id} ──────────────────────", "info")

            body = get_message_body(msg)
            if not body.strip():
                self._log("Corps vide — ignoré.", "warn")
                manual_count += 1
                continue

            demandes = parse_body(body)
            if not demandes:
                self._log("Aucune demande parsée dans ce mail — À traiter manuellement.", "warn")
                manual_count += 1
                continue

            all_ok = True
            for d in demandes:
                self._log(f"  Demande : {d.prenom} {d.nom} <{d.email}> → {', '.join(d.brochures)}", "info")

                if not d.is_valid:
                    reason = " | ".join(d.errors) if d.errors else "Champs invalides"
                    self._log(f"  ⚠️  À traiter manuellement : {reason}", "warn")
                    manual_count += 1
                    all_ok = False
                    continue

                # Excel
                excel_ok = False
                try:
                    log_demande(
                        excel_path=cfg["excel_path"],
                        sheet_name=cfg["excel_sheet"],
                        nom=d.nom,
                        prenom=d.prenom,
                        email=d.email,
                        telephone=d.telephone,
                        scolarite=d.scolarite,
                        brochures=d.brochures,
                        statut="En cours" if not dry_run else "DRY-RUN",
                    )
                    excel_ok = True
                    self._log(f"  ✓ Excel enregistré.", "ok")
                except PermissionError as e:
                    self._log(f"  ✗ Excel : {e}", "err")
                    all_ok = False
                    error_count += 1
                except Exception as e:
                    self._log(f"  ✗ Excel erreur inattendue : {e}", "err")
                    all_ok = False
                    error_count += 1

                # Brevo
                brevo_result = send_brochures(
                    api_key=cfg["brevo_api_key"],
                    sender_name=cfg["sender_name"],
                    sender_email=cfg["sender_email"],
                    to_email=d.email,
                    prenom=d.prenom,
                    brochures=d.brochures,
                    brochure_files=cfg["brochure_files"],
                    brochures_dir=cfg["brochures_dir"],
                    antilles_list=cfg.get("antilles_brochures", []),
                    url_metropole=cfg.get("inscription_url_metropole", ""),
                    url_antilles=cfg.get("inscription_url_antilles", ""),
                    dry_run=dry_run,
                )

                if brevo_result["success"]:
                    tag = "warn" if brevo_result.get("dry_run") else "ok"
                    suffix = " (DRY-RUN)" if brevo_result.get("dry_run") else f" — id: {brevo_result.get('message_id','')}"
                    self._log(f"  ✓ Email envoyé{suffix}", tag)
                    # Mise à jour statut Excel
                    if excel_ok and not dry_run:
                        try:
                            _update_last_row_status(cfg["excel_path"], cfg["excel_sheet"], "Envoyé")
                        except Exception:
                            pass
                else:
                    self._log(f"  ✗ Brevo : {brevo_result['error']}", "err")
                    all_ok = False
                    error_count += 1

            if all_ok and not manual_count:
                try:
                    mark_as_processed(service, msg_id, label_id)
                    self._log(f"  ✓ Mail marqué traité.", "ok")
                    ok_count += 1
                except Exception as e:
                    self._log(f"  ✗ Impossible de marquer le mail : {e}", "err")

        summary = f"\nTerminé — OK: {ok_count}  |  Manuel: {manual_count}  |  Erreurs: {error_count}"
        self._log(summary, "ok" if error_count == 0 else "warn")
        self.after(0, lambda: self.status_lbl.config(text=summary.strip()))
        self._finish()

    def _finish(self):
        self.after(0, lambda: self.btn_process.config(
            state="normal", text="▶  Traiter les nouvelles demandes"
        ))


def _update_last_row_status(excel_path: str, sheet_name: str, statut: str):
    """Met à jour le statut de la dernière ligne ajoutée."""
    import openpyxl
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet_name]
    last_row = ws.max_row
    # Colonne Statut = 8
    ws.cell(row=last_row, column=8).value = statut
    wb.save(excel_path)


if __name__ == "__main__":
    app = App()
    app.mainloop()
